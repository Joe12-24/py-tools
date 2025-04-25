import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

def create_issue_tracker_template(output_path="问题跟踪系统.xlsx"):
    # 初始化工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 移除默认sheet

    # 基础数据配置
    base_columns = [
        "编号", "提交日期", "问题类型", "优先级", "问题描述",
        "责任人", "解决方案", "解决状态", "解决日期", 
        "验证状态", "验证日期", "工时估算", "实际工时", "备注"
    ]
    
    # 数据验证规则
    type_dv = DataValidation(type="list", formula1='"系统BUG,功能优化,新增需求,数据维护"', allow_blank=True)
    priority_dv = DataValidation(type="list", formula1='"高,中,低"', allow_blank=True)
    status_dv = DataValidation(type="list", formula1='"未解决,已解决,延期处理"', allow_blank=True)
    verify_dv = DataValidation(type="list", formula1='"未验证,已验证,验证失败"', allow_blank=True)

    # 创建12个月份的sheet
    for month in range(1, 13):
        # 创建月度sheet
        sheet_name = f"2024-{month:02d}"
        ws = wb.create_sheet(title=sheet_name)
        
        # 添加表头
        ws.append(base_columns)
        
        # 设置数据验证
        ws.add_data_validation(type_dv)
        ws.add_data_validation(priority_dv)
        ws.add_data_validation(status_dv)
        ws.add_data_validation(verify_dv)
        
        # 应用数据验证到列
        for row in range(2, 100):  # 预置100行数据区
            type_dv.add(ws[f"C{row}"])     # 问题类型列
            priority_dv.add(ws[f"D{row}"])  # 优先级列
            status_dv.add(ws[f"H{row}"])    # 解决状态列
            verify_dv.add(ws[f"J{row}"])    # 验证状态列

        # 添加统计公式
        stats_row = 102  # 统计行位置
        stats = [
            ("BUG类总工时",        "=SUMIFS(L2:L100,C2:C100,\"系统BUG\")"),
            ("新增需求总工时",      "=SUMIFS(L2:L100,C2:C100,\"新增需求\")*1.2"),
            ("平均解决周期(天)",    "=AVERAGE(IF(ISNUMBER(I2:I100), I2:I100-TODAY()))"),
            ("验证通过率",         "=COUNTIF(J2:J100,\"已验证\")/COUNTA(J2:J100)"),
            ("总工时",            "=SUM(L2:L100)"), ]
        
        for idx, (label, formula) in enumerate(stats, start=1):
            ws[f"A{stats_row+idx}"] = label
            ws[f"B{stats_row+idx}"] = formula

        # 设置列宽
        col_widths = [8, 12, 12, 8, 40, 10, 30, 10, 12, 10, 12, 10, 10, 20]
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64+idx)].width = width

    # 创建年度汇总表
    summary_ws = wb.create_sheet(title="年度汇总")
    summary_data = {
        "月份": [f"2024-{m:02d}" for m in range(1,13)],
        "问题总数": [f'=COUNTIF(INDIRECT("{"2024-%02d" % m}!A:A"),"<>")-1' for m in range(1,13)],
        "BUG数量": [f'=COUNTIF(INDIRECT("{"2024-%02d" % m}!C:C"),"系统BUG")' for m in range(1,13)],
        "总工时": [f'=INDIRECT("{"2024-%02d" % m}!B107")' for m in range(1,13)]
    }
    df = pd.DataFrame(summary_data)
    for r in dataframe_to_rows(df, index=False, header=True):
        summary_ws.append(r)

    # 保存文件
    wb.save(output_path)
    return os.path.abspath(output_path)

# 生成文件（在当前目录生成）
file_path = create_issue_tracker_template()
print(f"文件已生成：{file_path}")